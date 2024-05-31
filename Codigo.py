import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

# Função para obter um número float válido do usuário
def obter_numero_float(mensagem):
    while True:
        try:
            valor = float(input(mensagem).replace(",", "."))
            if valor <= 0:
                raise ValueError("O valor deve ser um número positivo.")
            return valor
        except ValueError:
            print("Por favor, digite um número positivo válido.")

# Função para obter o tipo de despesa (variável ou fixa) do usuário
def obter_tipo_despesa(mensagem):
    while True:
        tipo = input(mensagem).upper()
        if tipo in ['V', 'F']:
            return tipo
        else:
            print("Por favor, digite 'V' para variável ou 'F'.")

# Pergunta ao usuário sobre o salário
salario = obter_numero_float("Digite o seu salário: R$")

# Inicializa dicionários para despesas e tipo de despesa (variável ou fixa)
despesas = {}
tipo_despesa = {}
ponderacoes = {}

# Define ponderações para as categorias
ponderacoes_default = {
     "água": 0.3,     # Menos reduzido
    "gás": 0.3,      # Menos reduzido
    "internet": 1.0,
    "luz": 0.7,      # Menos reduzido
    "mercado": 1.0,
    "urgência": 0.2,
    "lazer": 0.9,    # Mais reduzido
    "faculdade": 1.0,
    "outros": 0.5
}

# Pergunta ao usuário sobre as despesas
categorias = ["água", "gás", "internet", "luz", "mercado", "urgência", "lazer", "faculdade", "outros"]
for categoria in categorias:
    valor = obter_numero_float(f"Digite o valor da despesa de {categoria}: R$")
    tipo = obter_tipo_despesa(f"A despesa de {categoria} é variável (V) ou fixa (F)? ")
    despesas[categoria] = valor
    tipo_despesa[categoria] = tipo
    ponderacoes[categoria] = ponderacoes_default[categoria]

# Calcula o total gasto no mês
total_gasto = sum(despesas.values())

# Verifica se está gastando mais do que ganha
if total_gasto > salario:
    print("\nVocê está gastando mais do que ganha.")
else:
    print("\nVocê está gastando dentro do seu orçamento.")

# Pergunta ao usuário quanto ele gostaria de poupar
valor_poupanca = obter_numero_float("Digite o valor que deseja poupar: R$")

# Calcula quanto precisa ser reduzido das despesas variáveis
valor_a_reduzir = (total_gasto + valor_poupanca) - salario

# Calcula a redução necessária em cada categoria variável
if valor_a_reduzir > 0:
    despesas_reduzidas = despesas.copy()
    despesas_variaveis = {categoria: valor for categoria, valor in despesas.items() if tipo_despesa[categoria] == 'V'}
    total_variaveis = sum(despesas_variaveis.values())

    if total_variaveis > 0:
        soma_ponderacoes = sum(ponderacoes[categoria] for categoria in despesas_variaveis)
        fator_reducao = valor_a_reduzir / soma_ponderacoes

        for categoria in despesas_variaveis:
            reducao = fator_reducao * ponderacoes[categoria]
            # Assegura que a despesa não será reduzida abaixo de um valor mínimo
            despesas_reduzidas[categoria] = max(despesas_variaveis[categoria] - reducao, despesas_variaveis[categoria] * 0.01)
        total_gasto_reduzido = sum(despesas_reduzidas.values())
    else:
        print("Não há despesas variáveis para ajustar.")
        total_gasto_reduzido = total_gasto
else:
    despesas_reduzidas = despesas.copy()
    total_gasto_reduzido = total_gasto

# Informações sobre a redução
if valor_a_reduzir > 0:
    print(f"\nPara atingir o objetivo de poupar R${valor_poupanca:.2f}, você precisa reduzir suas despesas variáveis em R${valor_a_reduzir:.2f}.")
    print("Despesas reduzidas:")
    for categoria, valor in despesas_reduzidas.items():
        if tipo_despesa[categoria] == 'V':
            porcentagem_original = (despesas[categoria] / salario) * 100
            porcentagem_reduzida = (valor / salario) * 100
            reducao_percentual = ((despesas[categoria] - valor) / despesas[categoria]) * 100
            print(f"{categoria}:")
            print(f" - Original: R${despesas[categoria]:.2f} ({porcentagem_original:.2f}% do salário)")
            print(f" - Reduzido: R${valor:.2f} ({porcentagem_reduzida:.2f}% do salário, redução de {reducao_percentual:.2f}%)")
else:
    print(f"\nVocê está dentro do orçamento e ainda pode poupar R${valor_poupanca:.2f}.")

# Cria uma planilha com os resultados
df = pd.DataFrame({
    "Categoria": list(despesas.keys()) + ["Total"],
    "Valor Original (R$)": [val for val in list(despesas.values())] + [total_gasto],
    "Valor Reduzido (R$)": [val for val in list(despesas_reduzidas.values())] + [total_gasto_reduzido],
    "Porcentagem Original (%)": [(valor / salario) * 100 for valor in list(despesas.values())] + [(total_gasto / salario) * 100],
    "Porcentagem Reduzida (%)": [(valor / salario) * 100 for valor in list(despesas_reduzidas.values())] + [(total_gasto_reduzido / salario) * 100],
    "Redução (%)": [((despesas[categoria] - val) / despesas[categoria]) * 100 if tipo_despesa[categoria] == 'V' else np.nan for categoria, val in despesas_reduzidas.items()] + [np.nan]
})

# Salva os dados em um arquivo Excel
arquivo_excel = "despesas.xlsx"
df.to_excel(arquivo_excel, index=False)

# Carrega a planilha para aplicar formatações e adicionar gráficos
wb = load_workbook(arquivo_excel)
ws = wb.active

# Aplica estilos aos cabeçalhos
header_font = Font(bold=True)
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Ajusta a largura das colunas
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col)
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Adiciona gráficos à planilha
# Gráfico de Pizza - Antes da Redução
pie_chart = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=len(despesas) + 1)
data = Reference(ws, min_col=2, min_row=1, max_row=len(despesas) + 1)
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(labels)
pie_chart.title = "Distribuição de Despesas Original"

ws.add_chart(pie_chart, "H5")

# Gráfico de Pizza - Depois da Redução
pie_chart_reduzido = PieChart()
data_reduzida = Reference(ws, min_col=3, min_row=1, max_row=len(despesas) + 1)
pie_chart_reduzido.add_data(data_reduzida, titles_from_data=True)
pie_chart_reduzido.set_categories(labels)
pie_chart_reduzido.title = "Distribuição de Despesas Reduzida"

ws.add_chart(pie_chart_reduzido, "H20")

# Salva a planilha atualizada
wb.save(arquivo_excel)
print(f"\nOs resultados foram salvos no arquivo {arquivo_excel}.")
